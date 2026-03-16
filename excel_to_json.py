"""
Read 'барои барномаи нав.xlsx': columns П/Н (1st) and Номер счета (2nd).
Output:
- pension_numbers_numeric.json: rows where П/Н is numeric (asterisk stripped)
- pension_numbers_text.json: rows where П/Н is text
- pension_and_account_arrays.json: two arrays — pension_numbers and account_numbers (from numeric rows)
"""
import json
import re
from pathlib import Path

import openpyxl

# Paths: Excel and outputs in same dir as script
OUT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = OUT_DIR / "барои барномаи нав.xlsx"

def strip_asterisk(val):
    if val is None:
        return None
    s = str(val).strip()
    return s.rstrip("*").strip() if s else None

def is_numeric_pension(val):
    if val is None or val == "":
        return False
    s = str(val).strip()
    return s.isdigit() or (s.startswith("-") and s[1:].isdigit())

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices by header (first row)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_pn = headers.index("П/Н")
    except ValueError:
        col_pn = 0
    try:
        col_account = headers.index("Номер счета")
    except ValueError:
        col_account = 1

    numeric_rows = []
    text_rows = []

    for row in ws.iter_rows(min_row=2):
        raw_pn = row[col_pn].value
        raw_account = row[col_account].value

        pn_cleaned = strip_asterisk(raw_pn)
        # Account: keep as string; if Excel gave a float (long number), convert via int to avoid scientific notation
        if raw_account is None:
            account_str = ""
        elif isinstance(raw_account, (int, float)):
            account_str = str(int(raw_account))
        else:
            account_str = str(raw_account).strip()

        if pn_cleaned is None or pn_cleaned == "":
            continue
        if not account_str or account_str == "None":
            continue

        if is_numeric_pension(pn_cleaned):
            numeric_rows.append({
                "pension_number": int(pn_cleaned),
                "account_number": account_str,
            })
        else:
            text_rows.append({
                "pension_number": pn_cleaned,
                "account_number": account_str,
            })

    wb.close()

    out_numeric = OUT_DIR / "pension_numbers_numeric.json"
    out_text = OUT_DIR / "pension_numbers_text.json"

    with open(out_numeric, "w", encoding="utf-8") as f:
        json.dump(numeric_rows, f, ensure_ascii=False, indent=2)

    with open(out_text, "w", encoding="utf-8") as f:
        json.dump(text_rows, f, ensure_ascii=False, indent=2)

    # Two arrays: pension_numbers and account_numbers (from numeric rows)
    out_arrays = OUT_DIR / "pension_and_account_arrays.json"
    arrays = {
        "pension_numbers": [r["pension_number"] for r in numeric_rows],
        "account_numbers": [r["account_number"] for r in numeric_rows],
    }
    with open(out_arrays, "w", encoding="utf-8") as f:
        json.dump(arrays, f, ensure_ascii=False, indent=2)

    print(f"Written {len(numeric_rows)} numeric rows -> {out_numeric}")
    print(f"Written {len(text_rows)} text rows -> {out_text}")
    print(f"Written arrays -> {out_arrays}")

if __name__ == "__main__":
    main()